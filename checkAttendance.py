#coding:utf-8

import xlrd
import xlwt
from openpyxl import load_workbook
import time
from datetime import date,datetime


#the colx of department
depColx=0
#our departments
deps=[u'安全软件测试与支持部',u'安管标准化产品及平台部',
u'安管行业化产品部',u'安管互联网产品部',u'安全软件技术支持部',
u'大数据分析与安全研究部']

#openpyxl will spend too much time
def testOpenpyxl(filename):
    start=time.time()
    wb2=load_workbook(filename)
    print wb2.get_sheet_names() 
    print 'openpyxl: ',time.time()-start

#xlrd is much more efficent
def testXlrd(filename):
    start=time.time()
    #, encoding_override="gbk"
    book=xlrd.open_workbook(filename)
    sh=book.sheet_by_index(0)
    print "Worksheet name(s): ",book.sheet_names()[0]
    #print 'xlrd: ', time.time()-start
    print book.biff_version
    print book.codepage
    print book.encoding
    #sh=book.sheet_by_index(0)
    #print sh.name, sh.nrows, sh.ncols
    #print "Cell D30 is", sh.cell_value(rowx=29, colx=3)
    #for rx in range(sh.nrows):
    #    print sh.row(rx)
    
def testXlwt(filename):
    book=xlwt.Workbook()
    sheet1=book.add_sheet('hello')
    book.add_sheet('world')
    sheet1.write(0,0,'hello')
    sheet1.write(0,1,'world')
    row1 = sheet1.row(1)
    row1.write(0,'A2')
    row1.write(1,'B2')
    
    sheet1.col(0).width = 10000
    
    sheet2 = book.get_sheet(1)
    sheet2.row(0).write(0,'Sheet 2 A1')
    sheet2.row(0).write(1,'Sheet 2 B1')
    sheet2.flush_row_data()
    
    sheet2.write(1,0,'Sheet 2 A3')
    sheet2.col(0).width = 5000
    sheet2.col(0).hidden = True
    
    book.save(filename)
    #book.save(TemporaryFile())

def pickMember(fileRead,fileWrite):
    book=xlrd.open_workbook(fileRead)
    shr=book.sheet_by_index(0)
    newBook=xlwt.Workbook()
    shw=newBook.add_sheet('sheet1')
    #print 'rows: ', shr.nrows
    #print 'columns: ', shr.ncols
    #print 'A15: ', shr.cell_value(rowx=14,colx=0).encode('gbk')
    
    cnt=0
    for rx in range(shr.nrows):
        #print 'A'+str(rx)+': ', sh.cell_value(rowx=rx,colx=depColx).encode('gbk')
        dep=shr.cell_value(rowx=rx,colx=depColx)
        if dep in deps:
            #print dep.encode('gbk')
            for cx in range(shr.ncols):
                #print shr.cell_value(rowx=rx,colx=depColx).encode('gbk')
                if cx==2:
                    tmp=shr.cell_value(rx,cx)
                    try:
                        #print tmp
                        dateValue=xlrd.xldate_as_tuple(tmp,book.datemode)
                        print date(*dateValue[:3])
                        shw.write(cnt,cx,date(*dateValue[:3]))
                    except Exception,e:
                        print shw.write(cnt,cx,shr.cell_value(rx,cx))
                else:
                    shw.write(cnt,cx,shr.cell_value(rx,cx))
            cnt+=1
    newBook.save(fileWrite)
    
                
                
if __name__=='__main__':
    #testXlrd('June.xls')
    #testOpenpyxl('June.xlsx')
    #testXlwt('xlwt.xls')
    pickMember('June.xls','treated.xls')

